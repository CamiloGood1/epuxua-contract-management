"""
EPUXUA — Script de limpieza de datos del Excel
Corrige los 10 problemas de calidad antes de migrar a Supabase.
Salida: carpeta clean_data/ con CSVs listos para importar.

Uso:
  pip install openpyxl pandas unidecode
  python EPUXUA_CLEAN.py
"""

import re
import os
import json
import uuid
import math
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd
from unidecode import unidecode

# ── Configuración ─────────────────────────────────────────────

EXCEL_PATH = Path.home() / "Downloads" / "Contratación Epuxua E.I.C.E.xlsx"
OUT_DIR    = Path(__file__).parent / "clean_data"
OUT_DIR.mkdir(exist_ok=True)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PROBLEMA 1 — Estados inconsistentes (27+ variantes → 9 valores normalizados)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

STATUS_MAP = {
    # EN_EJECUCION
    "EN EJECUCIÓN": "EN_EJECUCION",
    "EN EJECUCION": "EN_EJECUCION",
    "EN EJEUCIÓN":  "EN_EJECUCION",
    "EN EJUCUCIÓN": "EN_EJECUCION",
    "EN EJEUCUCIÓN": "EN_EJECUCION",
    "EN EJECUCIÓN \nPENDIENTE LIBERAR SALDO": "EN_EJECUCION",
    "EN EJECUCIÓN\nPENDIENTE ACTA DE LIQUIDACIÓN": "EN_EJECUCION",
    "EN EJUCUCIÓN": "EN_EJECUCION",
    # SUSPENDIDO
    "SUSPENDIDO":   "SUSPENDIDO",
    "SUSPENSIÓN":   "SUSPENDIDO",
    # TERMINADO
    "TERMINADO":    "TERMINADO",
    "TERMINADO\nPENDIENTE ACTA DE LIQUIDACIÓN":    "TERMINADO",
    "TERMINADO\nCIERRE PROYECTADO PENDIENTE POLIZA": "TERMINADO",
    "TERMINADO\nREVISAR CIERRE":                   "TERMINADO",
    "TERMINADO ANORMALMENTE":                       "TERMINADO",
    "PROCESO TERMINADO":                            "TERMINADO",
    # TERMINADO_ANTICIPADAMENTE
    "TERMINADO ANTICIPADAMENTE":                    "TERMINADO_ANTICIPADAMENTE",
    "TERMINADO\nANTICIPADAMENTE":                   "TERMINADO_ANTICIPADAMENTE",
    "TERMINACIÓN ANTICIPADA":                       "TERMINADO_ANTICIPADAMENTE",
    "TERMINADO ANTICIADAMENTE\nCIERRE CONTRACTUAL": "TERMINADO_ANTICIPADAMENTE",
    "TERMINADO ANTICIPADAMENTE\nSIN LIQUIDACIÓN":   "TERMINADO_ANTICIPADAMENTE",
    "TEMINADO ANTICIPADAMENTE Y LIQUIDADO":         "TERMINADO_ANTICIPADAMENTE",
    "TERMINADO ANTICIPADAMENTE\nSIN LIQUIDACION":   "TERMINADO_ANTICIPADAMENTE",
    # LIQUIDADO
    "LIQUIDADO":                        "LIQUIDADO",
    "LIQUIDADO ":                       "LIQUIDADO",
    "TERMINADO Y LIQUIDADO ANTICIPADAMENTE": "LIQUIDADO",
    "LIQUIDADO PENDIENTE CARGUE SECOP": "LIQUIDADO",
    # CIERRE_CONTRACTUAL
    "CIERRE CONTRACTUAL":               "CIERRE_CONTRACTUAL",
    # DECLARADO_FALLIDO
    "DECLARADO FALLIDO":                "DECLARADO_FALLIDO",
    # ACTA_NO_EJECUCION
    "ACTA DE NO EJECUCIÓN":             "ACTA_NO_EJECUCION",
    "ACTA DE NO EJECUCIÓN \nPENDIENTE": "ACTA_NO_EJECUCION",
    # NO_SUSCRIPCION
    "NO SUSCRIPCIÓN":                   "NO_SUSCRIPCION",
    # CERRADO (solo interadmin)
    "CERRADO":                          "CIERRE_CONTRACTUAL",
    # Sin datos
    "#REF!":   None,
    "":        None,
    "None":    None,
}

def normalize_status(raw: Optional[str]) -> Optional[str]:
    if raw is None:
        return None
    clean = str(raw).strip().upper()
    # Buscar match exacto
    if clean in STATUS_MAP:
        return STATUS_MAP[clean]
    # Buscar por clave normalizada
    for k, v in STATUS_MAP.items():
        if k.upper() == clean:
            return v
    # Fallback: si contiene "EN EJECUCI"
    if "EN EJECUCI" in clean:
        return "EN_EJECUCION"
    if "SUSPENDID" in clean:
        return "SUSPENDIDO"
    if "TERMINADO ANTICIPAD" in clean:
        return "TERMINADO_ANTICIPADAMENTE"
    if "LIQUIDADO" in clean:
        return "LIQUIDADO"
    if "TERMINADO" in clean:
        return "TERMINADO"
    if "CIERRE" in clean:
        return "CIERRE_CONTRACTUAL"
    print(f"  ⚠️  Estado no mapeado: '{raw}'")
    return None

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PROBLEMA 2 — Modalidades inconsistentes (15+ variantes → enum)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

MODALITY_MAP = {
    "CONTRATACION DIRECTA":               "CONTRATACION_DIRECTA",
    "CONTRATACIÓN DIRECTA":               "CONTRATACION_DIRECTA",
    "CONTRATACION DIRECTA ":              "CONTRATACION_DIRECTA",
    "PAGO CONTRA FACTURA":                "PAGO_FACTURA",
    "INVITACIÓN ABIERTA":                 "INVITACION_ABIERTA",
    "INVITACION ABIERTA":                 "INVITACION_ABIERTA",
    "INVITACIÓN A PRESELECCIONADOS":      "INVITACION_PRESELECCIONADOS",
    "INVITACION A PRESELECCIONADOS":      "INVITACION_PRESELECCIONADOS",
    "INVITACIÓN A OFERENTES PRESELECCIONADOS":  "INVITACION_PRESELECCIONADOS",
    "INVITACIÓN A OFERENTES PRESELECIONADOS":   "INVITACION_PRESELECCIONADOS",
    "INVITACION A OFERENTES PRESELECCIONADOS":  "INVITACION_PRESELECCIONADOS",
    "INVITACIÓN A OFERENTES PRESELECIONADOS ":  "INVITACION_PRESELECCIONADOS",
    "CONCURSO DE MERITOS":                "CONCURSO_MERITOS",
    "CONCURSO DE MÉRITOS":                "CONCURSO_MERITOS",
    "ORDEN DE COMPRA":                    "ORDEN_COMPRA",
    "TIENDA VIRTUAL":                     "TIENDA_VIRTUAL",
    "ACUERDO MARCO":                      "ACUERDO_MARCO",
}

def normalize_modality(raw: Optional[str]) -> str:
    if not raw:
        return "CONTRATACION_DIRECTA"
    clean = str(raw).strip().upper()
    if clean in MODALITY_MAP:
        return MODALITY_MAP[clean]
    for k, v in MODALITY_MAP.items():
        if k.upper() == clean:
            return v
    if "DIRECTA" in clean or "DIRECTA" in unidecode(clean):
        return "CONTRATACION_DIRECTA"
    if "PRESELECCIONADO" in clean:
        return "INVITACION_PRESELECCIONADOS"
    if "ABIERTA" in clean:
        return "INVITACION_ABIERTA"
    if "TIENDA" in clean:
        return "TIENDA_VIRTUAL"
    if "COMPRA" in clean:
        return "ORDEN_COMPRA"
    print(f"  ⚠️  Modalidad no mapeada: '{raw}'")
    return "CONTRATACION_DIRECTA"

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PROBLEMA 3 — Normalización de texto (nombres, áreas)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def norm_text(s: Optional[str]) -> str:
    """Minúsculas, sin tildes, espacios internos colapsados — para comparación.
    ' '.join(...split()) colapsa dobles espacios igual que title_case_name,
    garantizando que 'FABIAN  CAMILO' y 'FABIAN CAMILO' sean el mismo key."""
    if not s:
        return ""
    return unidecode(" ".join(str(s).strip().split()).lower())

def title_case_name(s: Optional[str]) -> str:
    """Juan Pablo Cañon → normaliza casing."""
    if not s:
        return ""
    return " ".join(w.capitalize() for w in str(s).strip().split())

AREA_MAP = {
    "gerencia":                                              "Gerencia",
    "secretaria general":                                    "Secretaría General",
    "secretaria general ":                                   "Secretaría General",
    "subgerencia administrativa, financiera y gestion humana": "Subgerencia Administrativa, Financiera y Gestión Humana",
    "subgerencia administrativa financiera y de gestion humana": "Subgerencia Administrativa, Financiera y Gestión Humana",
    "subgerencia administrativa, financiera y de gestion humana": "Subgerencia Administrativa, Financiera y Gestión Humana",
    "subgerencia administrativa, financiera y gestion h":   "Subgerencia Administrativa, Financiera y Gestión Humana",
    "subgerencia":                                          "Subgerencia Administrativa, Financiera y Gestión Humana",
    "subgerencia tecnica":                                  "Subgerencia Técnica",
    "subgerente tecnico":                                   "Subgerencia Técnica",
    "direccion de estrcuturacion de proyectos":             "Dirección de Estructuración de Proyectos",
    "direccion estructuracion de proyectos":                "Dirección de Estructuración de Proyectos",
    "subgerenca tecnica":                                   "Subgerencia Técnica",
    "direccion de operacion urbana":                        "Dirección de Operación Urbana",
    "direccion de ejecucion de proyectos":                  "Dirección de Ejecución de Proyectos",
    "direccion de estructuracion de proyectos":             "Dirección de Estructuración de Proyectos",
    "direccion de estructuracion de proyecto":              "Dirección de Estructuración de Proyectos",
    "direccion de servicios publicos":                      "Dirección de Servicios Públicos",
}

def normalize_area(raw: Optional[str]) -> Optional[str]:
    if not raw:
        return None
    key = norm_text(raw)
    if key in AREA_MAP:
        return AREA_MAP[key]
    # Búsqueda parcial
    for k, v in AREA_MAP.items():
        if k in key or key in k:
            return v
    print(f"  ⚠️  Área no mapeada: '{raw}'")
    return raw.strip()

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PROBLEMA 4 — Prórrogas en texto libre
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

_MONTHS = {"mes": 30, "meses": 30, "month": 30}
_DAYS   = {"dia": 1, "días": 1, "dias": 1, "day": 1}
_YEARS  = {"año": 365, "años": 365, "year": 365}

def _split_presupuesto_tokens(raw: Optional[str], max_len: int = 50) -> list[str]:
    if not raw or str(raw).strip().upper() in ("N/A", "NA", "NONE", ""):
        return []
    text = str(raw).replace("\r\n", "\n").replace("\r", "\n")
    tokens: list[str] = []
    for line in text.split("\n"):
        line = line.strip()
        if not line:
            continue
        for part in re.split(r"[\s,;|]+", line):
            p = part.strip()
            if p and p.upper() not in ("N/A", "NA", "NONE"):
                tokens.append(p[:max_len])
    return tokens

def _expand_budget_commitment_rows(base: dict) -> list[dict]:
    """Varios CDP/CRP en una celda → una fila por número (cabecera en varchar(50))."""
    nums = _split_presupuesto_tokens(base.get("number"), 50)
    codes = _split_presupuesto_tokens(base.get("budget_code"), 120)
    if len(nums) <= 1:
        if nums:
            base["number"] = nums[0]
        if codes:
            base["budget_code"] = codes[0]
        return [base]
    try:
        val = float(base.get("value") or 0)
    except (TypeError, ValueError):
        val = 0.0
    rows = []
    for i, num in enumerate(nums):
        row = dict(base)
        row["id"] = base["id"] if i == 0 else str(uuid.uuid4())
        row["number"] = num
        row["budget_code"] = codes[i] if i < len(codes) else (codes[0] if codes else None)
        row["value"] = val if i == 0 else 0
        rows.append(row)
    return rows

def parse_term_days(text: Optional[str]) -> Optional[int]:
    """
    Intenta convertir un plazo en texto a días.
    Ej: "4 MESES" → 120, "65 dias" → 65, "4 meses\n3 meses" → 210 (suma)
    Retorna None si no puede parsear.
    """
    if not text or str(text).strip().upper() in ("N/A", "NA", "NONE", ""):
        return None
    total = 0
    found = False
    # Buscar patrones como "4 MESES", "12 dias", "1 AÑO"
    pattern = re.compile(
        r'(\d+(?:[.,]\d+)?)\s*(mes(?:es)?|d[ií]as?|a[ñn]os?|months?|days?|years?)',
        re.IGNORECASE
    )
    for match in pattern.finditer(str(text)):
        num  = float(match.group(1).replace(",", "."))
        unit = unidecode(match.group(2).lower().strip())
        if any(unit.startswith(k) for k in _MONTHS):
            total += round(num * 30)
            found = True
        elif any(unit.startswith(k) for k in _DAYS):
            total += round(num)
            found = True
        elif any(unit.startswith(k) for k in _YEARS):
            total += round(num * 365)
            found = True
    return total if found else None

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PROBLEMA 5 — Fechas con serial inválido de Excel
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def clean_date(val) -> Optional[str]:
    """Convierte fecha de Excel a string ISO 'YYYY-MM-DD'. None si inválida."""
    if val is None:
        return None
    if isinstance(val, datetime):
        # datetime primero (es subclase de date, evaluar antes)
        try:
            d = val.date()
            if date(2000, 1, 1) <= d <= date(2040, 12, 31):
                return d.isoformat()
        except Exception:
            pass
        return None
    if isinstance(val, date):
        try:
            if date(2000, 1, 1) <= val <= date(2040, 12, 31):
                return val.isoformat()
        except Exception:
            pass
        return None
    if isinstance(val, str):
        val_clean = val.strip()
        if val_clean.upper() in ("N/A", "NA", "", "NONE"):
            return None
        # Intentar parsear formatos comunes
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                d = datetime.strptime(val_clean, fmt).date()
                if date(2000, 1, 1) <= d <= date(2040, 12, 31):
                    return d.isoformat()
            except ValueError:
                continue
    # Número flotante (serial Excel) — ignorar si ya falló openpyxl
    if isinstance(val, (int, float)) and not math.isnan(val):
        # Serial válido de Excel: 1 = 1900-01-01
        if 36526 <= val <= 54787:  # 2000-2050 aprox
            try:
                from openpyxl.utils.datetime import from_excel
                d = from_excel(val)
                if isinstance(d, (datetime, date)):
                    d2 = d if isinstance(d, date) else d.date()
                    if date(2000, 1, 1) <= d2 <= date(2040, 12, 31):
                        return d2.isoformat()
            except Exception:
                pass
    return None

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PROBLEMA 6 — Valores numéricos con formato moneda
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def clean_numeric(val) -> Optional[float]:
    """Limpia valores como '$1.234.567', '1,234,567.89', '0' → float."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if math.isnan(val):
            return None
        return float(val)
    s = str(val).strip()
    if s.upper() in ("N/A", "NA", "", "NONE", "0"):
        return 0.0
    # Quitar símbolo de moneda y espacios
    s = re.sub(r"[$ ]", "", s)
    # Manejar separadores colombianos: 1.234.567,89 → 1234567.89
    if re.match(r"^\d{1,3}(\.\d{3})+,\d+$", s):
        s = s.replace(".", "").replace(",", ".")
    elif re.match(r"^\d{1,3}(\.\d{3})+$", s):
        s = s.replace(".", "")
    elif "," in s:
        s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PROBLEMA 7 — Deduplicación de supervisores por nombre
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class SupervisorCatalog:
    """
    Catálogo deduplicado de supervisores.
    La clave de deduplicación es el nombre normalizado (unidecode + lower).
    Múltiples grafías del mismo nombre → un solo registro.
    """
    def __init__(self):
        self._by_norm: dict[str, dict] = {}

    def get_or_create(self, raw_name: Optional[str]) -> Optional[str]:
        """Retorna el UUID del supervisor, creándolo si no existe."""
        if not raw_name or str(raw_name).strip().upper() in ("N/A", "NA", "NONE", ""):
            return None
        key = norm_text(raw_name)
        if not key:
            return None
        if key not in self._by_norm:
            # Primer nombre visto = nombre canónico (title case)
            self._by_norm[key] = {
                "id":        str(uuid.uuid4()),
                "full_name": title_case_name(raw_name),
                "variants":  [raw_name.strip()],
            }
        else:
            # Registrar variante adicional para el log
            variant = raw_name.strip()
            if variant not in self._by_norm[key]["variants"]:
                self._by_norm[key]["variants"].append(variant)
        return self._by_norm[key]["id"]

    def to_df(self) -> pd.DataFrame:
        rows = [
            {"id": v["id"], "full_name": v["full_name"],
             "active": True, "variants": "; ".join(v["variants"])}
            for v in self._by_norm.values()
        ]
        return pd.DataFrame(rows)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PROBLEMA 9 — Contratistas sin CC/NIT (dedup por nombre)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class ContractorCatalog:
    def __init__(self):
        self._by_doc:  dict[tuple, dict] = {}  # (doc_number, doc_type) → entry
        self._by_norm: dict[str, dict]   = {}  # norm_name → entry

    def get_or_create(
        self,
        name: Optional[str],
        person_type: Optional[str],
        doc_number: Optional[str] = None,
        doc_type:   Optional[str] = None,
    ) -> Optional[str]:
        if not name or str(name).strip().upper() in ("N/A", "NA", "NONE", ""):
            return None

        clean_name   = title_case_name(name)
        norm_name    = norm_text(name)
        clean_pt     = "NATURAL" if str(person_type or "").upper() in ("NATURAL","N") else "JURIDICA"
        clean_doc    = str(doc_number).strip() if doc_number else None
        clean_dt     = doc_type.upper() if doc_type else None

        # Buscar por documento (más preciso)
        if clean_doc:
            key_doc = (clean_doc, clean_dt or "CC")
            if key_doc in self._by_doc:
                return self._by_doc[key_doc]["id"]

        # Buscar por nombre normalizado
        if norm_name in self._by_norm:
            entry = self._by_norm[norm_name]
            # Si ahora tenemos el documento y antes no lo teníamos, actualizar
            if clean_doc and not entry.get("document_number"):
                entry["document_number"] = clean_doc
                entry["document_type"]   = clean_dt or "CC"
                self._by_doc[(clean_doc, clean_dt or "CC")] = entry
            return entry["id"]

        # Crear nuevo
        entry = {
            "id":              str(uuid.uuid4()),
            "full_name":       clean_name,
            "person_type":     clean_pt,
            "document_number": clean_doc,
            "document_type":   clean_dt,
        }
        self._by_norm[norm_name] = entry
        if clean_doc:
            self._by_doc[(clean_doc, clean_dt or "CC")] = entry
        return entry["id"]

    def to_df(self) -> pd.DataFrame:
        seen = set()
        rows = []
        for entry in list(self._by_norm.values()) + list(self._by_doc.values()):
            eid = entry["id"]
            if eid in seen:
                continue
            seen.add(eid)
            rows.append({
                "id":              entry["id"],
                "full_name":       entry["full_name"],
                "person_type":     entry["person_type"],
                "document_number": entry.get("document_number"),
                "document_type":   entry.get("document_type"),
                "active":          True,
            })
        return pd.DataFrame(rows)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CARGA Y LIMPIEZA POR HOJA
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

print("Cargando Excel…")
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

supervisors  = SupervisorCatalog()
contractors  = ContractorCatalog()

# Catálogo de áreas (se crea en BD con datos iniciales, aquí solo mapeamos)
area_catalog: dict[str, str] = {}  # norm → UUID asignado en runtime

def area_id(raw: Optional[str]) -> Optional[str]:
    """Retorna nombre normalizado del área (FK se resuelve en migración SQL)."""
    return normalize_area(raw)

# ── Contratos por año ─────────────────────────────────────────

YEAR_SHEETS = {
    2021: ("Contratación_2021",   {"tipo": "DIRECTO",
        "cols": {
            "recurso": 0, "numero": 1, "proceso": 2, "modalidad": 3,
            "contratista": 4, "objeto": 5, "persona": 6, "clase": 7,
            "area": 8, "supervisor": 9, "suscripcion": 10, "plazo_txt": 11,
            "inicio": 12, "valor_inicial": 13, "adicion": 14,
            "prorroga_txt": 16, "terminacion": 17, "pagado": 18,
            "vigencia": 20, "rubro": 21, "cdp": 22, "fecha_cdp": 23,
            "crp": 24, "fecha_crp": 25, "observaciones": 26, "estado": 27,
            "link": None,
        }
    }),
    2022: ("Contratación_2022",   {"tipo": "DIRECTO",
        "cols": {
            "recurso": 0, "numero": 1, "proceso": 2, "modalidad": 3,
            "contratista": 4, "objeto": 5, "persona": 6, "clase": 7,
            "area": 8, "supervisor": 9, "suscripcion": 10, "plazo_txt": 11,
            "inicio": 12, "valor_inicial": 13, "adicion": 14,
            "prorroga_txt": 16, "terminacion": 17, "pagado": 18,
            "vigencia": 20, "rubro": 20, "cdp": 21, "fecha_cdp": 22,
            "crp": 23, "fecha_crp": 24, "observaciones": 27, "estado": 28,
            "link": 30,
        }
    }),
    2023: ("Contratación_2023",   {"tipo": "DIRECTO",
        "cols": {
            "recurso": 0, "numero": 1, "proceso": 2, "modalidad": 3,
            "contratista": 4, "objeto": 5, "persona": 6, "clase": 7,
            "area": 8, "supervisor": 9, "suscripcion": 10, "plazo_txt": 11,
            "inicio": 12, "valor_inicial": 13, "adicion": 14,
            "prorroga_txt": 16, "terminacion": 17, "pagado": 18,
            "vigencia": 20, "rubro": 22, "cdp": 23, "fecha_cdp": 24,
            "crp": 25, "fecha_crp": 26, "observaciones": 27, "estado": 28,
            "link": 33,
        }
    }),
    2024: ("Contratación _2024",  {"tipo": "DIRECTO",
        "cols": {
            "recurso": 0, "numero": 1, "proceso": 2, "modalidad": 3,
            "contratista": 4, "objeto": 5, "persona": 6, "clase": 7,
            "area": 8, "supervisor": 9, "suscripcion": 10, "plazo_txt": 11,
            "inicio": 12, "valor_inicial": 13, "adicion": 14,
            "prorroga_txt": 16, "terminacion": 17, "pagado": 18,
            "vigencia": 20, "rubro": 22, "cdp": 23, "fecha_cdp": 24,
            "crp": 25, "fecha_crp": 26, "observaciones": 27, "estado": 28,
            "link": None,
        }
    }),
    2025: ("Contratación_2025",   {"tipo": "DIRECTO",
        "cols": {
            "recurso": 0, "numero": 1, "proceso": 2, "modalidad": 3,
            "contratista": 4, "objeto": 5, "persona": 6, "clase": 7,
            "area": 8, "supervisor": 9, "publicacion": 10, "suscripcion": 11,
            "plazo_txt": 12, "inicio": 13, "valor_mensual": 14,
            "valor_inicial": 15, "adicion": 16, "terminacion": 29,
            "pagado": 30, "vigencia": None, "rubro": 33,
            "cdp": 34, "fecha_cdp": 35, "crp": 36, "fecha_crp": 37,
            "observaciones": 42, "estado": 43,
            "fecha_liquidacion": 44, "fecha_cierre": 45,
            "poliza_numero": 48, "poliza_aprobacion": 49, "poliza_entidad": 50,
            "interventor": 51, "link": None,
        }
    }),
    2026: ("Contratacion_2026",   {"tipo": "DIRECTO",
        "cols": {
            "recurso": 0, "numero": 1, "proceso": 2, "modalidad": 3,
            "contratista": 4, "doc_contratista": 5, "objeto": 6,
            "persona": 7, "clase": 8, "area": 9, "supervisor": 10,
            "doc_supervisor": 11, "secop_url": 12, "publicacion": 13,
            "suscripcion": 14, "plazo_txt": 15, "inicio": 16,
            "poliza_numero": 17, "poliza_entidad": 18, "poliza_expedicion": 19,
            "poliza_inicio": 20, "poliza_fin": 21, "poliza_aprobacion": 22,
            "valor_mensual": 23, "valor_inicial": 24, "recurso2": 25,
            "cdp": 26, "valor_cdp": 27, "fecha_cdp": 28, "rubro": 29,
            "crp": 30, "valor_crp": 31, "fecha_crp": 32,
            "tipo_modificacion": 33, "fecha_adicion": 34, "valor_adicion": 35,
            "cdp_adicion": 36, "valor_cdp_adicion": 37, "fecha_cdp_adicion": 38,
            "rubro_adicion": 39, "crp_adicion": 40, "valor_crp_adicion": 41,
            "fecha_crp_adicion": 42, "valor_final": 43,
            "fecha_prorroga": 44, "plazo_prorroga": 45,
            "terminacion": 49, "fecha_liquidacion": 50, "estado": 51,
            "mipymes_consultados": 52, "mipymes_presentados": 53,
            "proveedores_presentados": 54, "mipymes_pr": 55,
            "mipymes_beneficiados": 56, "mipymes_participaron": 57,
            "limitado_mipymes": 58, "adjudicado_mipymes": 59,
        }
    }),
}

def get_cell(row: tuple, col_idx: Optional[int], default=None):
    """Obtiene el valor de una celda por índice, con manejo de None."""
    if col_idx is None or col_idx >= len(row):
        return default
    val = row[col_idx]
    return val if val is not None else default

def is_interadmin_ref(val: Optional[str]) -> bool:
    """¿El campo PROYECTO apunta a un número de contrato interadmin?"""
    if not val:
        return False
    # Patrón: 4 dígitos - 4 dígitos (ej: 3407-2021, 4168-2022, 0499-2024)
    return bool(re.match(r"^\d{3,5}-\d{4}$", str(val).strip()))

# Errores conocidos en columna PROYECTO / objeto (Excel)
PARENT_REF_OVERRIDES = {
    "286-2024": "140-2024",   # IMRDS deportes; no existe interadmin 286
    "154-2024": "1395-2024",  # PROYECTO copió número del hijo; objeto cita 1395
    "2257-2025": "02257-2025",
}

def ref_variants(ref: str) -> list[str]:
    ref = re.sub(r"\s+", "-", str(ref).strip())
    out = [ref]
    m = re.match(r"^(\d+)-(\d{4})$", ref)
    if not m:
        return out
    num, yr = m.group(1), m.group(2)
    out.append(f"{int(num)}-{yr}")
    out.append(f"{num.zfill(5)}-{yr}")
    out.append(f"{num.zfill(4)}-{yr}")
    return list(dict.fromkeys(out))

def extract_interadmin_refs_from_text(text: Optional[str]) -> list[str]:
    if not text:
        return []
    refs = []
    for m in re.finditer(r"(\d{3,5})\s*[- ]\s*(\d{4})", str(text), re.I):
        refs.append(f"{m.group(1)}-{m.group(2)}")
    return refs

def resolve_parent_ref(
    ref: str,
    object_text: Optional[str],
    interadmin_numbers: set[str],
) -> Optional[str]:
    ref = str(ref).strip()
    candidates: list[str] = []
    if ref in PARENT_REF_OVERRIDES:
        candidates.append(PARENT_REF_OVERRIDES[ref])
    candidates.append(ref)
    for v in ref_variants(ref):
        candidates.append(v)
    for extracted in extract_interadmin_refs_from_text(object_text):
        candidates.extend([extracted, *ref_variants(extracted)])
    for c in candidates:
        c = re.sub(r"\s+", "-", c.strip())
        for v in ref_variants(c):
            if v in interadmin_numbers:
                return v
    return None

def fix_parent_contract_refs(df: pd.DataFrame) -> pd.DataFrame:
    interadmin_numbers = set(
        df.loc[df["contract_type"] == "INTERADMINISTRATIVO", "contract_number"]
        .astype(str).str.strip()
    )
    fixed = 0
    unresolved = []
    for idx, row in df.iterrows():
        pref = row.get("parent_contract_ref")
        if pref is None or (isinstance(pref, float) and pd.isna(pref)):
            continue
        pref = str(pref).strip()
        if pref in ("", "nan", "None"):
            continue
        resolved = resolve_parent_ref(pref, row.get("object"), interadmin_numbers)
        if resolved:
            if resolved != pref:
                df.at[idx, "parent_contract_ref"] = resolved
                fixed += 1
        else:
            unresolved.append((row["contract_number"], pref))
    if fixed:
        print(f"  parent_ref corregidos automáticamente: {fixed}")
    if unresolved:
        print(f"  parent_ref sin interadmin en Excel ({len(unresolved)}):")
        for cn, pr in unresolved:
            print(f"    · {cn} → {pr} (falta contrato INTERADMINISTRATIVO en fuente)")
    return df

# Acumuladores
contracts_rows       = []
amendments_rows      = []
extensions_rows      = []
cdp_crp_rows         = []
policies_rows        = []
mipymes_rows         = []
interadmin_rows      = []
invoice_detail_rows  = []

print("Procesando contratos por año…")

for year, (sheet_name, cfg) in YEAR_SHEETS.items():
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    cols = cfg["cols"]
    processed = 0
    skipped   = 0

    for row in rows[1:]:  # Saltar encabezado
        # ── PROBLEMA 7: Filas de relleno con ceros ──────────────
        numero_raw = get_cell(row, cols.get("numero"))
        if numero_raw is None:
            skipped += 1
            continue
        numero = str(numero_raw).strip()
        if numero in ("0", "", "None", "Contrato", "Número del Contrato",
                      "Número Del Contrato", "DECLARADO FALLIDO"):
            skipped += 1
            continue

        # Estado normalizado
        status = normalize_status(get_cell(row, cols.get("estado")))
        if status is None:
            status = "CIERRE_CONTRACTUAL"  # Fallback seguro

        # Personas
        supervisor_id  = supervisors.get_or_create(get_cell(row, cols.get("supervisor")))
        contractor_id  = contractors.get_or_create(
            name        = get_cell(row, cols.get("contratista")),
            person_type = get_cell(row, cols.get("persona")),
            doc_number  = get_cell(row, cols.get("doc_contratista")),
            doc_type    = "NIT" if str(get_cell(row, cols.get("persona"), "")).upper() == "JURIDICA" else "CC",
        )
        if contractor_id is None:
            skipped += 1
            continue

        # Recurso / proyecto padre
        recurso_raw = get_cell(row, cols.get("recurso"))
        parent_ref  = None
        resource_type = None
        if recurso_raw and is_interadmin_ref(recurso_raw):
            parent_ref = str(recurso_raw).strip()  # Se resolverá como FK en migración
        else:
            resource_type = str(recurso_raw).strip() if recurso_raw else None
            # Normalizar recurso
            if resource_type:
                resource_type = re.sub(r"\s+", " ", resource_type.upper().strip())
                if "FUNCIONAMIENTO" in resource_type:
                    resource_type = "FUNCIONAMIENTO"
                elif "OPERACIÓN" in resource_type or "OPERACION" in resource_type:
                    resource_type = "GASTO DE OPERACIÓN COMERCIAL"

        # Valores financieros
        valor_inicial = clean_numeric(get_cell(row, cols.get("valor_inicial"))) or 0.0
        adicion_val   = clean_numeric(get_cell(row, cols.get("adicion")))       or 0.0
        pagado        = clean_numeric(get_cell(row, cols.get("pagado")))         or 0.0
        vigencia      = clean_numeric(get_cell(row, cols.get("vigencia")))       or 0.0
        val_mensual   = clean_numeric(get_cell(row, cols.get("valor_mensual")))

        # Fechas
        fecha_sus  = clean_date(get_cell(row, cols.get("suscripcion")))
        if not fecha_sus:
            skipped += 1
            continue  # Fecha de suscripción es obligatoria

        fecha_ini  = clean_date(get_cell(row, cols.get("inicio")))
        fecha_fin  = clean_date(get_cell(row, cols.get("terminacion")))
        fecha_liq  = clean_date(get_cell(row, cols.get("fecha_liquidacion")))
        fecha_cie  = clean_date(get_cell(row, cols.get("fecha_cierre")))
        fecha_pub  = clean_date(get_cell(row, cols.get("publicacion")))

        # Plazo
        plazo_txt  = get_cell(row, cols.get("plazo_txt"))
        plazo_days = parse_term_days(plazo_txt)

        # Clase y área
        clase = get_cell(row, cols.get("clase"))
        if not clase:
            clase = "Prestación de Servicios"
        area_name = area_id(get_cell(row, cols.get("area")))

        # URLs
        secop_url = get_cell(row, cols.get("secop_url")) or get_cell(row, cols.get("observaciones"))
        if secop_url and not str(secop_url).startswith("http"):
            secop_url = None
        tech_url = get_cell(row, cols.get("link"))
        if tech_url and not str(tech_url).startswith("http"):
            tech_url = None

        contract_id = str(uuid.uuid4())

        contracts_rows.append({
            "id":                    contract_id,
            "contract_number":       numero,
            "selection_process_number": str(get_cell(row, cols.get("proceso")) or "").strip() or None,
            "year":                  year,
            "contract_type":         cfg["tipo"],
            "selection_modality":    normalize_modality(get_cell(row, cols.get("modalidad"))),
            "contract_class":        str(clase).strip(),
            "resource_type":         resource_type,
            "parent_contract_ref":   parent_ref,  # número de contrato padre; FK resuelta post-migración
            "contractor_id":         contractor_id,
            "supervisor_area":       area_name,   # nombre normalizado; FK resuelta en SQL
            "supervisor_name":       title_case_name(str(get_cell(row, cols.get("supervisor")) or "")),
            "supervisor_id":         supervisor_id,
            "object":                str(get_cell(row, cols.get("objeto")) or "").strip(),
            "subscription_date":     fecha_sus,
            "publication_date":      fecha_pub,
            "start_date":            fecha_ini,
            "initial_term_text":     str(plazo_txt or "").strip() or None,
            "initial_term_days":     plazo_days,
            "end_date":              fecha_fin,
            "liquidation_date":      fecha_liq,
            "file_closure_date":     fecha_cie,
            "monthly_value":         val_mensual,
            "initial_value":         valor_inicial,
            "total_additions_value": adicion_val,
            "paid_value":            pagado,
            "future_validity":       vigencia,
            "status":                status,
            "secop_url":             secop_url,
            "technical_file_url":    tech_url,
            "interventor":           str(get_cell(row, cols.get("interventor")) or "").strip() or None,
            "observations":          str(get_cell(row, cols.get("observaciones")) or "").strip() or None,
        })
        processed += 1

        # ── Adición → contract_amendments ────────────────────────
        if adicion_val and adicion_val != 0:
            amendments_rows.append({
                "id":              str(uuid.uuid4()),
                "contract_id":     contract_id,
                "amendment_number": 1,
                "modification_type": str(get_cell(row, cols.get("tipo_modificacion")) or "ADICION").upper(),
                "amendment_value": adicion_val,
                "amendment_date":  clean_date(get_cell(row, cols.get("fecha_adicion"))),
                "cdp_number":      str(get_cell(row, cols.get("cdp_adicion")) or "").strip() or None,
                "cdp_value":       clean_numeric(get_cell(row, cols.get("valor_cdp_adicion"))),
                "cdp_date":        clean_date(get_cell(row, cols.get("fecha_cdp_adicion"))),
                "cdp_budget_code": str(get_cell(row, cols.get("rubro_adicion")) or "").strip() or None,
                "crp_number":      str(get_cell(row, cols.get("crp_adicion")) or "").strip() or None,
                "crp_value":       clean_numeric(get_cell(row, cols.get("valor_crp_adicion"))),
                "crp_date":        clean_date(get_cell(row, cols.get("fecha_crp_adicion"))),
                "crp_budget_code": None,
            })

        # ── CDP/CRP → budget_commitments ─────────────────────────
        cdp_num = str(get_cell(row, cols.get("cdp")) or "").strip()
        if cdp_num and cdp_num not in ("N/A", "NA", "None", ""):
            cdp_crp_rows.extend(_expand_budget_commitment_rows({
                "id":              str(uuid.uuid4()),
                "contract_id":     contract_id,
                "commitment_type": "CDP",
                "number":          cdp_num,
                "value":           clean_numeric(get_cell(row, cols.get("valor_cdp"))) or valor_inicial,
                "budget_code":     str(get_cell(row, cols.get("rubro")) or "").strip() or None,
                "date":            clean_date(get_cell(row, cols.get("fecha_cdp"))) or fecha_sus,
                "is_addition":     False,
            }))
        crp_num = str(get_cell(row, cols.get("crp")) or "").strip()
        if crp_num and crp_num not in ("N/A", "NA", "None", ""):
            cdp_crp_rows.extend(_expand_budget_commitment_rows({
                "id":              str(uuid.uuid4()),
                "contract_id":     contract_id,
                "commitment_type": "CRP",
                "number":          crp_num,
                "value":           clean_numeric(get_cell(row, cols.get("valor_crp"))) or valor_inicial,
                "budget_code":     str(get_cell(row, cols.get("rubro")) or "").strip() or None,
                "date":            clean_date(get_cell(row, cols.get("fecha_crp"))) or fecha_sus,
                "is_addition":     False,
            }))

        # ── Póliza → contract_policies ───────────────────────────
        poliza_num = str(get_cell(row, cols.get("poliza_numero")) or "").strip()
        if poliza_num and poliza_num not in ("N/A", "NA", ""):
            policies_rows.append({
                "id":             str(uuid.uuid4()),
                "contract_id":    contract_id,
                "policy_number":  poliza_num,
                "issuing_entity": str(get_cell(row, cols.get("poliza_entidad")) or "").strip() or None,
                "issue_date":     clean_date(get_cell(row, cols.get("poliza_expedicion"))),
                "start_date":     clean_date(get_cell(row, cols.get("poliza_inicio"))),
                "end_date":       clean_date(get_cell(row, cols.get("poliza_fin"))),
                "approval_date":  clean_date(get_cell(row, cols.get("poliza_aprobacion"))),
            })

        # ── MiPymes (2026) → mipymes_stats ───────────────────────
        if year == 2026:
            def to_bool(v):
                if v is None: return None
                return str(v).strip().lower() in ("si", "sí", "yes", "true", "1")
            mipymes_rows.append({
                "id":                  str(uuid.uuid4()),
                "contract_id":         contract_id,
                "providers_consulted": get_cell(row, cols.get("mipymes_consultados")),
                "mipymes_consulted":   get_cell(row, cols.get("mipymes_consultados")),
                "providers_presented": get_cell(row, cols.get("proveedores_presentados")),
                "mipymes_presented":   get_cell(row, cols.get("mipymes_pr")),
                "mipymes_benefited":   get_cell(row, cols.get("mipymes_beneficiados")),
                "mipymes_participated": to_bool(get_cell(row, cols.get("mipymes_participaron"))),
                "limited_to_mipymes":   to_bool(get_cell(row, cols.get("limitado_mipymes"))),
                "awarded_to_mipymes":   to_bool(get_cell(row, cols.get("adjudicado_mipymes"))),
            })

        # ── Prórroga texto libre (2021-2024) → contract_extensions
        prorroga_raw = get_cell(row, cols.get("prorroga_txt"))
        if prorroga_raw and str(prorroga_raw).strip().upper() not in ("N/A", "NA", "NONE", ""):
            # Dividir por salto de línea (puede haber varias)
            partes = re.split(r"\n|\|", str(prorroga_raw))
            for i, parte in enumerate(partes, 1):
                parte = parte.strip()
                if not parte or parte.upper() in ("N/A", "NA"):
                    continue
                days = parse_term_days(parte)
                extensions_rows.append({
                    "id":                 str(uuid.uuid4()),
                    "contract_id":        contract_id,
                    "extension_number":   i,
                    "extension_term_text": parte,
                    "extension_term_days": days,
                    "extension_date":     None,
                    "new_end_date":       None,
                })

    print(f"  {year}: {processed} contratos procesados, {skipped} filas ignoradas")

# ── Contratos Interadministrativos ────────────────────────────
print("Procesando contratos interadministrativos…")
ws_int = wb['Contratos Interadministrativos']
rows_int = list(ws_int.iter_rows(values_only=True))
interadmin_processed = 0

for row in rows_int[1:]:
    numero_raw = get_cell(row, 0)
    if not numero_raw or str(numero_raw).strip() in ("", "None"):
        continue
    numero = str(numero_raw).strip()
    if not re.match(r"^\d{3,5}-\d{4}$", numero):
        continue

    status = normalize_status(get_cell(row, 25))
    if status is None:
        status = "CIERRE_CONTRACTUAL"

    supervisor_id = supervisors.get_or_create(get_cell(row, 6))
    # Contratista no aplica para interadmin (es la Secretaría)
    # Usamos un contratista genérico que representará a la entidad
    secretaria = str(get_cell(row, 2) or "").strip()
    contractor_id = contractors.get_or_create(
        name=secretaria, person_type="JURIDICA"
    )
    if not contractor_id:
        continue

    valor_inicial = clean_numeric(get_cell(row, 11)) or 0.0
    adicion_val   = clean_numeric(get_cell(row, 12)) or 0.0
    admin_fee_ini = clean_numeric(get_cell(row, 14)) or 0.0
    admin_fee_add = clean_numeric(get_cell(row, 15)) or 0.0
    mandato_ini   = clean_numeric(get_cell(row, 17)) or 0.0
    mandato_add   = clean_numeric(get_cell(row, 18)) or 0.0
    pending_col   = clean_numeric(get_cell(row, 20)) or 0.0
    vigencia      = clean_numeric(get_cell(row, 21)) or 0.0

    fecha_sus = clean_date(get_cell(row, 8))
    if not fecha_sus:
        continue

    contract_id = str(uuid.uuid4())

    # Extract year from contract number (ej: 3407-2021 → 2021)
    year_match = re.search(r"-(\d{4})$", numero)
    contract_year = int(year_match.group(1)) if year_match else 2021

    contracts_rows.append({
        "id":                    contract_id,
        "contract_number":       numero,
        "selection_process_number": None,
        "year":                  contract_year,
        "contract_type":         "INTERADMINISTRATIVO",
        "selection_modality":    normalize_modality(get_cell(row, 1)),
        "contract_class":        "Contrato Interadministrativo",
        "resource_type":         None,
        "parent_contract_ref":   None,
        "contractor_id":         contractor_id,
        "supervisor_area":       area_id(get_cell(row, 5)),
        "supervisor_name":       title_case_name(str(get_cell(row, 6) or "")),
        "supervisor_id":         supervisor_id,
        "object":                str(get_cell(row, 3) or "").strip(),
        "subscription_date":     fecha_sus,
        "publication_date":      None,
        "start_date":            clean_date(get_cell(row, 9)),
        "initial_term_text":     str(get_cell(row, 7) or "").strip() or None,
        "initial_term_days":     parse_term_days(get_cell(row, 7)),
        "end_date":              clean_date(get_cell(row, 24)),
        "liquidation_date":      None,
        "file_closure_date":     None,
        "monthly_value":         None,
        "initial_value":         valor_inicial,
        "total_additions_value": adicion_val,
        "paid_value":            0.0,
        "future_validity":       vigencia,
        "status":                status,
        "secop_url":             None,
        "technical_file_url":    None,
        "interventor":           None,
        "observations":          str(get_cell(row, 26) or "").strip() or None,
    })

    interadmin_rows.append({
        "id":                    str(uuid.uuid4()),
        "contract_id":           contract_id,
        "secretaria":            secretaria,
        "admin_fee_initial":     admin_fee_ini,
        "admin_fee_additions":   admin_fee_add,
        "mandate_pool_initial":  mandato_ini,
        "mandate_pool_additions": mandato_add,
        "pending_collection":    pending_col,
    })

    if adicion_val != 0:
        amendments_rows.append({
            "id":              str(uuid.uuid4()),
            "contract_id":     contract_id,
            "amendment_number": 1,
            "modification_type": "ADICION",
            "amendment_value": adicion_val,
            "amendment_date":  None,
            "cdp_number": None, "cdp_value": None, "cdp_date": None, "cdp_budget_code": None,
            "crp_number": None, "crp_value": None, "crp_date": None, "crp_budget_code": None,
        })
    interadmin_processed += 1

print(f"  Interadmin: {interadmin_processed} procesados")

# ── Pago contra Factura ───────────────────────────────────────
print("Procesando Pago contra Factura…")
ws_pcf = wb['Pago contra Factura']
rows_pcf = list(ws_pcf.iter_rows(values_only=True))
pcf_processed = 0

for row in rows_pcf[2:]:
    pf_num = get_cell(row, 0)
    if not pf_num or not str(pf_num).startswith("PF"):
        continue
    pf_num = str(pf_num).strip()

    # Extraer año del número PF-001-2024
    year_match = re.search(r"-(\d{4})$", pf_num)
    pcf_year = int(year_match.group(1)) if year_match else 2024

    contractor_id = contractors.get_or_create(
        name=get_cell(row, 2), person_type=get_cell(row, 5) or "NATURAL"
    )
    if not contractor_id:
        continue

    # Columnas verificadas directamente en el Excel:
    # [2] Proveedor  [3] Responsable Solicitud (casi siempre NULL)
    # [4] Objeto     [5] Persona Natural/Jurídica  [6] Area  [7] Supervisor
    # [8] N° comité  [9] Fecha aprobación comité  [10] Fecha facturación
    # [11] Valor pagado  [12] Estado  [13] Observaciones  [14] Valor Total
    supervisor_id = supervisors.get_or_create(get_cell(row, 7))
    total_val = clean_numeric(get_cell(row, 14)) or 0.0
    paid_val  = clean_numeric(get_cell(row, 11)) or 0.0

    status_raw = get_cell(row, 12)
    pcf_status_map = {
        "PAGO": "CIERRE_CONTRACTUAL", "EJECUTADO": "CIERRE_CONTRACTUAL",
        "PENDIENTE DE PAGO": "EN_EJECUCION", "PENDIENTE FACTURA": "EN_EJECUCION",
    }
    status = pcf_status_map.get(str(status_raw or "").upper().strip(), "EN_EJECUCION")

    fecha_factura = clean_date(get_cell(row, 10))   # [10] = Fecha de Facturación
    contract_id = str(uuid.uuid4())

    objeto = str(get_cell(row, 4) or "").strip()    # [4] = Objeto (corregido)
    if not objeto:
        objeto = str(get_cell(row, 2) or "").strip()  # fallback: nombre del proveedor

    contracts_rows.append({
        "id":                    contract_id,
        "contract_number":       pf_num,
        "selection_process_number": None,
        "year":                  pcf_year,
        "contract_type":         "PAGO_FACTURA",
        "selection_modality":    "PAGO_FACTURA",
        "contract_class":        "Pago contra Factura",
        "resource_type":         None,
        "parent_contract_ref":   None,
        "contractor_id":         contractor_id,
        "supervisor_area":       area_id(get_cell(row, 6)),
        "supervisor_name":       title_case_name(str(get_cell(row, 7) or "")),
        "supervisor_id":         supervisor_id,
        "object":                objeto,
        "subscription_date":     fecha_factura or f"{pcf_year}-01-01",
        "publication_date":      None,
        "start_date":            fecha_factura,
        "initial_term_text":     None,
        "initial_term_days":     None,
        "end_date":              clean_date(get_cell(row, 9)),   # [9] = fecha aprobación
        "liquidation_date":      None,
        "file_closure_date":     None,
        "monthly_value":         None,
        "initial_value":         total_val,
        "total_additions_value": 0.0,
        "paid_value":            paid_val,
        "future_validity":       0.0,
        "status":                status,
        "secop_url":             None,
        "technical_file_url":    None,
        "interventor":           None,
        "observations":          str(get_cell(row, 13) or "").strip() or None,
    })

    invoice_detail_rows.append({
        "id":                  str(uuid.uuid4()),
        "contract_id":         contract_id,
        "committee_number":    str(get_cell(row, 8) or "").strip() or None,
        "committee_act_info":  None,
        "invoice_date":        fecha_factura,
        "requesting_officer":  str(get_cell(row, 3) or "").strip() or None,  # [3] = Responsable Solicitud
    })
    pcf_processed += 1

print(f"  PCF: {pcf_processed} procesados")

# ── Tienda Virtual ────────────────────────────────────────────
print("Procesando Tienda Virtual…")
ws_tv = wb['Tienda_Virtual_2024-2021']
rows_tv = list(ws_tv.iter_rows(values_only=True))
tv_processed = 0

for row in rows_tv[1:]:
    order_num = get_cell(row, 0)
    if not order_num or not isinstance(order_num, (int, float)):
        continue
    order_num_str = str(int(order_num))

    # Año de la fecha de suscripción
    fecha_sus = clean_date(get_cell(row, 8))
    tv_year = int(fecha_sus[:4]) if fecha_sus else 2022

    contractor_id = contractors.get_or_create(
        name=get_cell(row, 2), person_type=get_cell(row, 4) or "JURIDICA"
    )
    if not contractor_id:
        continue

    supervisor_id = supervisors.get_or_create(get_cell(row, 7))
    valor_inicial = clean_numeric(get_cell(row, 11)) or 0.0
    adicion_val   = clean_numeric(get_cell(row, 12)) or 0.0
    pagado        = clean_numeric(get_cell(row, 15)) or 0.0
    status = normalize_status(get_cell(row, 18)) or "CIERRE_CONTRACTUAL"

    contract_id = str(uuid.uuid4())

    contracts_rows.append({
        "id":                    contract_id,
        "contract_number":       order_num_str,
        "selection_process_number": None,
        "year":                  tv_year,
        "contract_type":         "TIENDA_VIRTUAL",
        "selection_modality":    normalize_modality(get_cell(row, 1) or "Tienda Virtual"),
        "contract_class":        str(get_cell(row, 5) or "Acuerdo Marco").strip(),
        "resource_type":         None,
        "parent_contract_ref":   None,
        "contractor_id":         contractor_id,
        "supervisor_area":       area_id(get_cell(row, 6)),
        "supervisor_name":       title_case_name(str(get_cell(row, 7) or "")),
        "supervisor_id":         supervisor_id,
        "object":                str(get_cell(row, 3) or "").strip(),
        "subscription_date":     fecha_sus or f"{tv_year}-01-01",
        "publication_date":      None,
        "start_date":            clean_date(get_cell(row, 10)),
        "initial_term_text":     str(get_cell(row, 9) or "").strip() or None,
        "initial_term_days":     None,
        "end_date":              clean_date(get_cell(row, 14)),
        "liquidation_date":      None,
        "file_closure_date":     None,
        "monthly_value":         None,
        "initial_value":         valor_inicial,
        "total_additions_value": adicion_val,
        "paid_value":            pagado,
        "future_validity":       clean_numeric(get_cell(row, 17)) or 0.0,
        "status":                status,
        "secop_url":             str(get_cell(row, 21) or "").strip() or None,
        "technical_file_url":    None,
        "interventor":           None,
        "observations":          str(get_cell(row, 19) or "").strip() or None,
    })

    if adicion_val != 0:
        amendments_rows.append({
            "id":              str(uuid.uuid4()),
            "contract_id":     contract_id,
            "amendment_number": 1,
            "modification_type": "ADICION",
            "amendment_value": adicion_val,
            "amendment_date":  None,
            "cdp_number": None, "cdp_value": None, "cdp_date": None, "cdp_budget_code": None,
            "crp_number": None, "crp_value": None, "crp_date": None, "crp_budget_code": None,
        })
    tv_processed += 1

print(f"  Tienda Virtual: {tv_processed} procesados")

# ── Pagos (PAGOS) ─────────────────────────────────────────────
print("Procesando pagos…")
ws_pay = wb['PAGOS']
rows_pay = list(ws_pay.iter_rows(values_only=True))

# Construir lookup: (contract_number, year_str) → contract_id
contract_lookup = {
    (r["contract_number"], str(r["year"])): r["id"]
    for r in contracts_rows
}

payments_rows = []
pay_processed = 0
pay_skipped   = 0

for row in rows_pay[1:]:
    # PROBLEMA 8: Filas de relleno con ceros
    contrato_raw = get_cell(row, 3)
    if not contrato_raw or str(contrato_raw).strip() in ("0", "", "None", "contrato"):
        pay_skipped += 1
        continue
    egreso   = get_cell(row, 1)
    year_raw = get_cell(row, 2)
    if not egreso or not year_raw:
        pay_skipped += 1
        continue
    valor_raw = get_cell(row, 7)
    if not valor_raw or str(valor_raw).strip() in ("0", "", "None"):
        pay_skipped += 1
        continue

    contrato_num = str(contrato_raw).strip()
    year_str     = str(int(float(str(year_raw))))
    contract_id  = contract_lookup.get((contrato_num, year_str))
    if not contract_id:
        pay_skipped += 1
        continue

    bruto      = clean_numeric(valor_raw)     or 0.0
    descuentos = clean_numeric(get_cell(row, 8)) or 0.0
    fecha_pago = clean_date(get_cell(row, 5))

    if not fecha_pago or bruto == 0:
        pay_skipped += 1
        continue

    payments_rows.append({
        "id":                   str(uuid.uuid4()),
        "contract_id":          contract_id,
        "payment_number":       int(re.sub(r"[^\d]", "", str(egreso)) or "0"),
        "payment_date":         fecha_pago,
        "gross_value":          bruto,
        "deductions":           descuentos,
        "cumulative_percentage": clean_numeric(get_cell(row, 10)),
        "drive_url":            str(get_cell(row, 14) or "").strip() or None,
    })
    pay_processed += 1

print(f"  Pagos: {pay_processed} importados, {pay_skipped} ignorados (relleno/ceros)")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# EXPORTAR CSVs LIMPIOS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

print("\nExportando CSVs limpios…")

df_supervisors = supervisors.to_df()
df_supervisors.to_csv(OUT_DIR / "supervisors.csv", index=False)
print(f"  supervisors.csv      → {len(df_supervisors)} registros")

df_contractors = contractors.to_df()
df_contractors.to_csv(OUT_DIR / "contractors.csv", index=False)
print(f"  contractors.csv      → {len(df_contractors)} registros")

df_contracts = pd.DataFrame(contracts_rows)
df_contracts = fix_parent_contract_refs(df_contracts)
df_contracts.to_csv(OUT_DIR / "contracts.csv", index=False)
print(f"  contracts.csv        → {len(df_contracts)} registros")

df_amendments = pd.DataFrame(amendments_rows)
df_amendments.to_csv(OUT_DIR / "contract_amendments.csv", index=False)
print(f"  contract_amendments.csv → {len(df_amendments)} registros")

df_extensions = pd.DataFrame(extensions_rows)
df_extensions.to_csv(OUT_DIR / "contract_extensions.csv", index=False)
print(f"  contract_extensions.csv → {len(df_extensions)} registros")

df_cdp_crp = pd.DataFrame(cdp_crp_rows)
df_cdp_crp.to_csv(OUT_DIR / "budget_commitments.csv", index=False)
print(f"  budget_commitments.csv → {len(df_cdp_crp)} registros")

df_policies = pd.DataFrame(policies_rows)
df_policies.to_csv(OUT_DIR / "contract_policies.csv", index=False)
print(f"  contract_policies.csv  → {len(df_policies)} registros")

df_interadmin = pd.DataFrame(interadmin_rows)
df_interadmin.to_csv(OUT_DIR / "interadmin_contract_details.csv", index=False)
print(f"  interadmin_details.csv → {len(df_interadmin)} registros")

df_invoice = pd.DataFrame(invoice_detail_rows)
df_invoice.to_csv(OUT_DIR / "invoice_payment_details.csv", index=False)
print(f"  invoice_payment_details.csv → {len(df_invoice)} registros")

df_mipymes = pd.DataFrame(mipymes_rows)
df_mipymes.to_csv(OUT_DIR / "mipymes_stats.csv", index=False)
print(f"  mipymes_stats.csv    → {len(df_mipymes)} registros")

df_payments = pd.DataFrame(payments_rows)
df_payments.to_csv(OUT_DIR / "payments.csv", index=False)
print(f"  payments.csv         → {len(df_payments)} registros")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# REPORTE DE CALIDAD
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

print("\n" + "="*60)
print("REPORTE DE CALIDAD")
print("="*60)
print(f"Contratos totales migrados:      {len(df_contracts)}")
print(f"  - Directos:                    {(df_contracts.contract_type=='DIRECTO').sum()}")
print(f"  - Interadministrativos:        {(df_contracts.contract_type=='INTERADMINISTRATIVO').sum()}")
print(f"  - Tienda Virtual:              {(df_contracts.contract_type=='TIENDA_VIRTUAL').sum()}")
print(f"  - Pago contra Factura:         {(df_contracts.contract_type=='PAGO_FACTURA').sum()}")
print(f"Contratistas únicos:             {len(df_contractors)}")
print(f"Supervisores únicos:             {len(df_supervisors)}")
print(f"Pagos migrados:                  {len(df_payments)}")
print(f"Adiciones:                       {len(df_amendments)}")
print(f"Prórrogas (texto parseado):      {len(df_extensions)}")
print(f"Compromisos CDP/CRP:             {len(df_cdp_crp)}")
print(f"Pólizas:                         {len(df_policies)}")
print()
print("Distribución de estados:")
for status, count in df_contracts.status.value_counts().items():
    print(f"  {status:<35} {count}")
print()
print(f"Contratos con parent_ref (derivados): {df_contracts.parent_contract_ref.notna().sum()}")
print(f"Fechas de suscripción nulas:          {df_contracts.subscription_date.isna().sum()}")
print(f"Contratos sin área asignada:          {df_contracts.supervisor_area.isna().sum()}")
print()
print("Supervisores con múltiples variantes detectadas:")
sup_df = df_supervisors[df_supervisors.variants.str.contains(";")]
print(f"  {len(sup_df)} supervisores con grafías distintas (ver supervisors.csv columna 'variants')")
print()

# Generar log de anomalías
anomalies = []
for _, row in df_contracts.iterrows():
    if pd.isna(row.get("end_date")):
        anomalies.append({"contract": row["contract_number"], "year": row["year"], "issue": "sin_fecha_fin"})
    if row.get("initial_value", 0) == 0:
        anomalies.append({"contract": row["contract_number"], "year": row["year"], "issue": "valor_cero"})

pd.DataFrame(anomalies).to_csv(OUT_DIR / "anomalies_log.csv", index=False)
print(f"Anomalías detectadas:             {len(anomalies)} (ver anomalies_log.csv)")
print()
print(f"✅ Archivos exportados en: {OUT_DIR}")
print()
print("SIGUIENTE PASO:")
print("1. Ejecutar EPUXUA_DDL.sql en Supabase (SQL Editor)")
print("2. Revisar supervisors.csv columna 'variants' para confirmar deduplicación")
print("3. Revisar anomalies_log.csv para contratos con datos faltantes")
print("4. Importar CSVs con: EPUXUA_IMPORT.sql (se generará)")
